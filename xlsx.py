"""
Workato: Python snippets by Workato — Execute Python code
==========================================================
Action name: Generate supplier XLSX template

INPUT FIELDS (define in Workato action config):
  - fields        (string)  ← JSON array of field objects
  - lookups       (string)  ← JSON array of lookup rows
  - client_name   (string)
  - variant_name  (string)

OUTPUT FIELDS (define in Workato action config):
  - file_content  (string)  ← base64-encoded XLSX
  - file_name     (string)
"""

import base64
import io
import json
import re
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, quote_sheetname

BOLD = Font(bold=True)


def main(input):
    # ── Parse JSON string inputs ──
    # Workato passes complex objects as JSON strings when mapped from
    # webhook list/array datapills. Parse them if they arrive as strings.
    fields = input.get('fields', [])
    if isinstance(fields, str):
        fields = json.loads(fields)

    lookups = input.get('lookups', [])
    if isinstance(lookups, str):
        lookups = json.loads(lookups)

    client_name  = input.get('client_name', 'Client')
    variant_name = input.get('variant_name', 'Default')

    print(f"Generating template: {client_name} / {variant_name}")
    print(f"Fields: {len(fields)}, Lookup rows: {len(lookups)}")

    # ── Generate the workbook ──
    xlsx_bytes = build_workbook(fields, lookups)

    file_name = f"Supplier_Template_{sanitize_filename(client_name)}_{sanitize_filename(variant_name)}.xlsx"

    print(f"Generated {file_name} ({len(xlsx_bytes)} bytes)")

    return {
        "file_content": base64.b64encode(xlsx_bytes).decode('utf-8'),
        "file_name": file_name
    }


# ── Helper functions (called from main) ──────────────────────────


def sanitize_filename(name):
    """Remove characters unsafe for filenames."""
    return re.sub(r'[^\w\s-]', '', str(name).strip()).replace(' ', '_')


def sanitize_range_name(name):
    """
    Mirror the GAS sanitizeRangeName() logic:
      1. Replace whitespace runs with _
      2. Strip non-alphanumeric/non-underscore chars
      3. Prepend _ if result starts with a digit
    """
    sanitized = re.sub(r'\s+', '_', str(name).strip())
    sanitized = re.sub(r'[^A-Za-z0-9_]', '', sanitized)
    if sanitized and sanitized[0].isdigit():
        sanitized = '_' + sanitized
    return sanitized


def build_workbook(fields, lookups):
    """Build the XLSX workbook and return raw bytes."""
    wb = Workbook()
    DATA_ROWS = 500

    # ── Supplier Data sheet ──
    ws = wb.active
    ws.title = "Supplier Data"

    # ── Hidden Data_Lookups sheet ──
    lookup_ws = wb.create_sheet("Data_Lookups")
    lookup_ws.sheet_state = 'hidden'

    lookup_col = 1  # next available column in Data_Lookups

    # Field-name and lookup-name → column index (1-based)
    field_col_map = {}
    for i, f in enumerate(fields):
        field_col_map[f['name']] = i + 1
        if f.get('lookup_name'):
            field_col_map[f['lookup_name']] = i + 1

    # ── Write headers ──
    for i, f in enumerate(fields):
        cell = ws.cell(row=1, column=i + 1, value=f['name'])
        cell.font = BOLD

    # ── Process each field ──
    for i, field in enumerate(fields):
        col = i + 1
        col_letter = get_column_letter(col)
        fmt = str(field.get('data_format') or '').strip().lower()

        # ── Dependent dropdown ──────────────────────────────
        if 'dependent' in fmt and field.get('lookup_name') and field.get('depends_on'):
            parent_col_idx = field_col_map.get(field['depends_on'])

            if not parent_col_idx:
                print(f"  WARN: parent '{field['depends_on']}' not in template, falling back to flat list")
            else:
                groups = group_lookups_by_parent(lookups, field['lookup_name'])

                if groups:
                    lookup_col = write_named_ranges(
                        wb, lookup_ws, groups, lookup_col
                    )

                    # INDIRECT formula — native Excel, no conversion layer
                    parent_letter = get_column_letter(parent_col_idx)
                    formula = f'INDIRECT(SUBSTITUTE({parent_letter}2," ","_"))'

                    dv = DataValidation(
                        type="list",
                        formula1=formula,
                        allow_blank=True
                    )
                    dv.error = f"Select a valid {field['name']}"
                    dv.errorTitle = "Invalid selection"
                    dv.showErrorMessage = True
                    dv.add(f"{col_letter}2:{col_letter}{1 + DATA_ROWS}")
                    ws.add_data_validation(dv)

                    print(f"  {field['name']}: dependent dropdown → parent col {parent_letter}")
                    continue

        # ── Regular dropdown ────────────────────────────────
        if 'dropdown' in fmt and field.get('lookup_name'):
            values = get_flat_lookup_values(lookups, field['lookup_name'])

            if values:
                # Write values to Data_Lookups
                lookup_ws.cell(row=1, column=lookup_col, value=field['lookup_name']).font = BOLD

                for r, val in enumerate(values, start=2):
                    lookup_ws.cell(row=r, column=lookup_col, value=val)

                # Validation referencing the lookup column
                lk_letter = get_column_letter(lookup_col)
                sheet_ref = quote_sheetname('Data_Lookups')
                ref = f"={sheet_ref}!${lk_letter}$2:${lk_letter}${1 + len(values)}"

                dv = DataValidation(type="list", formula1=ref, allow_blank=True)
                dv.add(f"{col_letter}2:{col_letter}{1 + DATA_ROWS}")
                ws.add_data_validation(dv)

                lookup_col += 1
                print(f"  {field['name']}: dropdown ({len(values)} values)")

            continue

        # ── Date validation ─────────────────────────────────
        if field.get('data_type') == 'date':
            dv = DataValidation(type="date", allow_blank=True)
            dv.add(f"{col_letter}2:{col_letter}{1 + DATA_ROWS}")
            ws.add_data_validation(dv)
            print(f"  {field['name']}: date validation")

    # ── Auto-fit column widths (approximate) ──
    for i, f in enumerate(fields):
        ws.column_dimensions[get_column_letter(i + 1)].width = max(len(f['name']) + 4, 15)

    # ── Serialize ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def group_lookups_by_parent(lookups, lookup_name):
    """
    Group lookup rows by parent_value for a given table.
    Returns dict: { "Finance": ["Accountant", "Controller", ...], ... }
    """
    groups = {}
    for row in lookups:
        table = str(row.get('table_name', '')).strip()
        label = str(row.get('label', '')).strip()
        parent = str(row.get('parent_value', '')).strip()
        is_active = row.get('is_active', False)

        # Handle string booleans from JSON
        if isinstance(is_active, str):
            is_active = is_active.strip().upper() in ('TRUE', '1')

        if table == lookup_name and is_active and label and parent:
            groups.setdefault(parent, []).append(label)

    return groups


def get_flat_lookup_values(lookups, lookup_name):
    """Get active labels for a non-dependent lookup table."""
    values = []
    for row in lookups:
        table = str(row.get('table_name', '')).strip()
        label = str(row.get('label', '')).strip()
        is_active = row.get('is_active', False)

        if isinstance(is_active, str):
            is_active = is_active.strip().upper() in ('TRUE', '1')

        if table == lookup_name and is_active and label:
            values.append(label)

    return values


def write_named_ranges(wb, lookup_ws, groups, start_col):
    """
    Write each parent group as a column in Data_Lookups and create
    a workbook-scoped named range for it.

    Returns the next available column index.
    """
    col = start_col

    for parent_name, values in groups.items():
        range_name = sanitize_range_name(parent_name)

        # Header
        lookup_ws.cell(row=1, column=col, value=parent_name).font = BOLD

        # Values
        for r, val in enumerate(values, start=2):
            lookup_ws.cell(row=r, column=col, value=val)

        # Named range (workbook scope)
        lk_letter = get_column_letter(col)
        sheet_ref = quote_sheetname('Data_Lookups')
        ref = f"{sheet_ref}!${lk_letter}$2:${lk_letter}${1 + len(values)}"

        defn = DefinedName(range_name, attr_text=ref)
        wb.defined_names.add(defn)

        print(f"    Named range '{range_name}' → {ref} ({len(values)} values)")
        col += 1

    return col
