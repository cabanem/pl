#!/usr/bin/env python3
"""
build_demo_artifacts.py — Hand-roll the two JSONs needed for the demo from
the MARS master config workbook.

Outputs (in /home/claude/):
    mars_variant.json           — for the validator snippet
    mars_canonical_model.json   — for the XLSX parser (VAL-01 step 9.2)

Both target the same chosen variant (default: Variant_2, since that mirrored
the earlier work; flip CHOSEN_VARIANT_COL to 6 for Variant_1 if preferred).
"""
import json
from openpyxl import load_workbook

WORKBOOK    = '/mnt/user-data/uploads/config_MARS.xlsx'
VARIANT_NAME = 'Variant_2'
VARIANT_INDEX = 2
CHOSEN_VARIANT_COL = 7  # column index in 6_variants for Variant_2 (6 for Variant_1)
EXPECTED_SHEET_NAME = 'Data'   # what the supplier XLSX's data sheet will be called

wb = load_workbook(WORKBOOK, data_only=True, read_only=True)

# ── 1. Lift 4_fields verbatim ────────────────────────────────────────────────
# The validator reads header at index 7, data from index 8. We keep the whole
# sheet so the structure matches the original variant JSON exactly. Empty
# trailing rows are stripped — they're harmless but bloat the JSON.

ws = wb['4_fields']
all_fields_rows = list(ws.iter_rows(values_only=True))
# Strip trailing rows where col 2 (Field name) is empty AFTER row 8
last_real_row = 7
for i, r in enumerate(all_fields_rows):
    if i >= 8 and r[2]:
        last_real_row = i
fields_sheet = [list(r) for r in all_fields_rows[:last_real_row + 1]]

# ── 2. Lift 5_lookups verbatim ───────────────────────────────────────────────
# Validator reads header at index 4, data from index 5. Same trim treatment.

ws = wb['5_lookups']
all_lookup_rows = list(ws.iter_rows(values_only=True))
last_real_row = 4
for i, r in enumerate(all_lookup_rows):
    if i >= 5 and r and r[0]:
        last_real_row = i
# Trim each row to 6 columns: Table name, Value, Label, Parent value,
# Record active?, Project specific?. MARS's sheet has 9 columns (3 trailing
# blanks); the validator's `table, value, _label, parent, active, _proj = r`
# unpack expects exactly 6.
lookups_sheet = [list(r)[:6] for r in all_lookup_rows[:last_real_row + 1]]

# ── 3. Build _field_visibility from 6_variants ──────────────────────────────
# 6_variants row 4 is the header. Field name in col 1, Variant_1 in col 6,
# Variant_2 in col 7. A True means visible in that variant.

ws = wb['6_variants']
variants_rows = list(ws.iter_rows(values_only=True))
field_visibility = {}
for r in variants_rows[5:]:
    if not r or not r[1]:
        continue
    field_name = r[1]
    visible = bool(r[CHOSEN_VARIANT_COL]) if len(r) > CHOSEN_VARIANT_COL else False
    field_visibility[field_name] = visible

# ── 4. Assemble the variant JSON ────────────────────────────────────────────

variant_json = {
    "_meta": {
        "variant_name":  VARIANT_NAME,
        "variant_index": VARIANT_INDEX,
        "client_name":   "MARS RC",
        "source":        "hand-rolled from config_MARS.xlsx for demo",
        "expected_sheet_name": EXPECTED_SHEET_NAME,
    },
    "_field_visibility": field_visibility,
    "4_fields":  fields_sheet,
    "5_lookups": lookups_sheet,
}

# ── 5. Build the minimal canonical model from visible fields ────────────────
# The XLSX parser reads only _meta.expected_sheet_name and cfg_fields[*].field_name.
# Pull field_name from 4_fields where _field_visibility[name] == True.

header = fields_sheet[7]
name_col_idx = header.index('Field name')
visible_field_names = []
for r in fields_sheet[8:]:
    fname = r[name_col_idx]
    if fname and field_visibility.get(fname, False):
        visible_field_names.append(fname)

canonical_model = {
    "_meta": {
        "expected_sheet_name": EXPECTED_SHEET_NAME,
        "variant_name":        VARIANT_NAME,
        "client_name":         "MARS RC",
        "source":              "hand-rolled from config_MARS.xlsx for demo",
    },
    "cfg_fields": [{"field_name": n} for n in visible_field_names],
}

# ── 6. Write both files ──────────────────────────────────────────────────────

with open('/home/claude/mars_variant.json', 'w') as f:
    json.dump(variant_json, f, indent=2, default=str)

with open('/home/claude/mars_canonical_model.json', 'w') as f:
    json.dump(canonical_model, f, indent=2, default=str)

print(f"Variant: {VARIANT_NAME}")
print(f"4_fields rows lifted: {len(fields_sheet)}")
print(f"5_lookups rows lifted: {len(lookups_sheet)}")
print(f"Total fields in visibility map: {len(field_visibility)}")
print(f"Visible fields in {VARIANT_NAME}: {len(visible_field_names)}")
print(f"\nVisible field names:")
for n in visible_field_names:
    print(f"  - {n}")
