import base64
import json
import re
import traceback
from io import BytesIO


# Header cells may carry a trailing " *" required marker. That marker is
# display chrome — it is NOT part of the field identity. The builder writes
# required headers as `field_name + " *"`, but the canonical model and the
# downstream validator key everything by the bare `field_name`. Strip it on
# the way in so the header check and the record keys both match.
#
# Transition note: this lives on the reader side on purpose. Templates already
# issued with the marker are in suppliers' hands and can't be recalled, so the
# reader must tolerate both the starred and unstarred header shapes.
_REQUIRED_MARKER = re.compile(r"\s*\*\s*$")


def _undecorate(header):
    if header is None:
        return ""
    return _REQUIRED_MARKER.sub("", str(header).strip())


def main(input):
    # Running diagnostics. Included in the `log` of EVERY return so any failure
    # carries context — Workato Python actions don't surface print()/stdout.
    diag = {}

    def err(code, message):
        return {
            "ok": False,
            "error_code": code,
            "error_message": message,
            "rows_json": "[]",
            "row_count": 0,
            "log": json.dumps(diag),
        }

    try:
        # ── Resolve input to XLSX bytes ───────────────────────────────────
        # Workato hands a file/binary datapill to Python as `bytes`, not a
        # base64 string. Normalize to bytes, then decode ONLY if the bytes are
        # base64 *text* — never base64-decode bytes that are already the file.
        raw = input.get("xlsx_base64", b"")
        raw_b = raw.encode("latin-1") if isinstance(raw, str) else raw
        diag.update({
            "raw_type": type(raw).__name__,
            "raw_len": len(raw_b),
            "raw_head_hex": raw_b[:8].hex(),   # 504b0304 = file, 55457344 = base64 text
        })

        if raw_b[:4] == b"PK\x03\x04":
            xlsx_bytes = raw_b                      # already the workbook
        else:
            try:
                xlsx_bytes = base64.b64decode(raw_b, validate=True)
            except Exception as e:
                diag["b64_error"] = str(e)
                return err("err_submission_unparseable",
                           f"Input is neither XLSX bytes nor valid base64: {e}")

        diag["resolved_len"] = len(xlsx_bytes)
        diag["resolved_head_hex"] = xlsx_bytes[:4].hex()
        diag["eocd_present"] = b"PK\x05\x06" in xlsx_bytes[-1024:]

        # The magic number is the only thing that actually means "this is XLSX".
        if xlsx_bytes[:4] != b"PK\x03\x04":
            return err("err_submission_unparseable",
                       f"Resolved input is not an XLSX ({len(xlsx_bytes)} bytes, "
                       f"head={xlsx_bytes[:4].hex()})")

        # ── Parse canonical model ─────────────────────────────────────────
        try:
            canonical_model = json.loads(input["canonical_model_json"])
        except Exception as e:
            return err("err_submission_unparseable",
                       f"Could not parse canonical_model_json: {e}")

        expected_sheet = canonical_model.get("_meta", {}).get(
            "expected_sheet_name", "Data Entry")
        # Tolerate malformed cfg_fields entries instead of raising KeyError.
        field_names = [
            f["field_name"]
            for f in canonical_model.get("cfg_fields", [])
            if isinstance(f, dict) and f.get("field_name")
        ]

        # Descriptions back the locked instruction banner the builder writes at
        # sheet row 2. We use them to recognize and skip that row on the way
        # back in WITHOUT depending on its position — old, banner-less templates
        # still in flight won't match and are treated as ordinary data. Keyed by
        # the bare field_name, same as the undecorated headers below.
        descriptions = {
            f["field_name"]: (f.get("description") or "").strip()
            for f in canonical_model.get("cfg_fields", [])
            if isinstance(f, dict) and f.get("field_name")
        }
        has_descriptions = any(descriptions.values())

        # ── Open workbook ─────────────────────────────────────────────────
        # read_only is ON: supplier files reach ~300 cols x 10k rows (~3M
        # cells), and the full in-memory cell model would risk OOM at that size
        # — a crash that bypasses the err() contract. The streaming reader holds
        # one row at a time. reset_dimensions() (below) stops us trusting a
        # declared <dimension> that may be wrong/missing — the "odd dimension /
        # ragged row" hazard that previously argued for read_only=False is
        # handled here, and any residual reader error is caught by the outer
        # try and returned as the contract error shape rather than crashing.
        # data_only resolves formulas to last-saved cached values.
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
        except Exception as e:
            return err("err_submission_unparseable", f"Could not open XLSX: {e}")

        diag["sheets"] = wb.sheetnames

        # ── Locate expected sheet ─────────────────────────────────────────
        if expected_sheet not in wb.sheetnames:
            wb.close()
            return err("err_submission_structurally_invalid",
                       f"Required sheet '{expected_sheet}' not found. "
                       f"Found: {wb.sheetnames}")
        ws = wb[expected_sheet]
        ws.reset_dimensions()  # iterate by actual data, not a possibly-wrong declared dimension

        # ── Single pass: first row is the header, the rest are data ───────
        # One streaming iterator; nothing is materialized.
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            wb.close()
            return err("err_submission_structurally_invalid", "Header row is empty")

        # Strip the " *" required marker so identifiers match the canonical
        # model's field_name — fixes both the missing-headers check and the
        # record keys consumed by validate_upload (which look up by field_name).
        headers = [_undecorate(h) for h in header_row]
        missing_fields = [fn for fn in field_names if fn not in headers]
        if missing_fields:
            wb.close()
            return err("err_submission_structurally_invalid",
                       f"Missing required headers: {missing_fields}")

        def _is_instruction_banner(row_tuple):
            # Treat the row as the builder's instruction banner iff every
            # non-empty cell equals the corresponding field's description.
            # Requires at least one real match, so an all-blank row isn't
            # mistaken for the banner (and an all-blank row would be dropped
            # as blank by _ingest regardless).
            if not has_descriptions:
                return False
            matched = False
            for i, cell in enumerate(row_tuple):
                if i >= len(headers) or not headers[i]:
                    continue
                cell_s = "" if cell is None else str(cell).strip()
                desc = descriptions.get(headers[i], "")
                if cell_s == "" and desc == "":
                    continue
                if cell_s != desc:
                    return False
                matched = True
            return matched

        rows = []

        def _ingest(row_tuple):
            record = {}
            for i, cell in enumerate(row_tuple):
                if i < len(headers) and headers[i]:
                    record[headers[i]] = cell
            # Keep only rows with at least one non-blank value.
            if any(v is not None and str(v).strip() != "" for v in record.values()):
                rows.append(record)

        # Peek the first data row to detect/skip the instruction banner WITHOUT
        # materializing the sheet — the peek touches exactly one row; everything
        # after streams through the iterator. A banner-less template just ingests
        # its first row normally.
        first = next(row_iter, None)
        banner_skipped = first is not None and _is_instruction_banner(first)
        if first is not None and not banner_skipped:
            _ingest(first)
        diag["instruction_banner_skipped"] = banner_skipped

        for row_tuple in row_iter:
            _ingest(row_tuple)

        wb.close()

        # default=str coerces datetime/date (openpyxl) so json.dumps won't raise;
        # the connector's check_data_type / check_data_format accept strings.
        rows_json = json.dumps(rows, default=str)
        diag["row_count"] = len(rows)
        return {
            "ok": True,
            "error_code": None,
            "error_message": None,
            "rows_json": rows_json,
            "row_count": len(rows),
            "log": json.dumps(diag),
        }

    except Exception as e:
        # Catch-all: the action ALWAYS returns the contract shape and never
        # leaks a raw traceback into the recipe. Full stack lands in `log`.
        diag["traceback"] = traceback.format_exc()
        return err("err_unexpected", f"Unhandled error during extraction: {e}")            "row_count": 0,
            "log": json.dumps(diag),
        }

    try:
        # ── Resolve input to XLSX bytes ───────────────────────────────────
        # Workato hands a file/binary datapill to Python as `bytes`, not a
        # base64 string. Normalize to bytes, then decode ONLY if the bytes are
        # base64 *text* — never base64-decode bytes that are already the file.
        raw = input.get("xlsx_base64", b"")
        raw_b = raw.encode("latin-1") if isinstance(raw, str) else raw
        diag.update({
            "raw_type": type(raw).__name__,
            "raw_len": len(raw_b),
            "raw_head_hex": raw_b[:8].hex(),   # 504b0304 = file, 55457344 = base64 text
        })

        if raw_b[:4] == b"PK\x03\x04":
            xlsx_bytes = raw_b                      # already the workbook
        else:
            try:
                xlsx_bytes = base64.b64decode(raw_b, validate=True)
            except Exception as e:
                diag["b64_error"] = str(e)
                return err("err_submission_unparseable",
                           f"Input is neither XLSX bytes nor valid base64: {e}")

        diag["resolved_len"] = len(xlsx_bytes)
        diag["resolved_head_hex"] = xlsx_bytes[:4].hex()
        diag["eocd_present"] = b"PK\x05\x06" in xlsx_bytes[-1024:]

        # The magic number is the only thing that actually means "this is XLSX".
        if xlsx_bytes[:4] != b"PK\x03\x04":
            return err("err_submission_unparseable",
                       f"Resolved input is not an XLSX ({len(xlsx_bytes)} bytes, "
                       f"head={xlsx_bytes[:4].hex()})")

        # ── Parse canonical model ─────────────────────────────────────────
        try:
            canonical_model = json.loads(input["canonical_model_json"])
        except Exception as e:
            return err("err_submission_unparseable",
                       f"Could not parse canonical_model_json: {e}")

        expected_sheet = canonical_model.get("_meta", {}).get(
            "expected_sheet_name", "Data Entry")
        # Tolerate malformed cfg_fields entries instead of raising KeyError.
        field_names = [
            f["field_name"]
            for f in canonical_model.get("cfg_fields", [])
            if isinstance(f, dict) and f.get("field_name")
        ]

        # Descriptions back the locked instruction banner the builder writes at
        # sheet row 2. We use them to recognize and skip that row on the way
        # back in WITHOUT depending on its position — old, banner-less templates
        # still in flight won't match and are treated as ordinary data. Keyed by
        # the bare field_name, same as the undecorated headers below.
        descriptions = {
            f["field_name"]: (f.get("description") or "").strip()
            for f in canonical_model.get("cfg_fields", [])
            if isinstance(f, dict) and f.get("field_name")
        }
        has_descriptions = any(descriptions.values())

        # ── Open workbook ─────────────────────────────────────────────────
        # read_only is intentionally OFF: its streaming reader is the usual
        # source of internal IndexErrors on files with an odd dimension or
        # ragged rows, and supplier submissions are small. data_only resolves
        # formulas to their last-saved computed values.
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
        except Exception as e:
            return err("err_submission_unparseable", f"Could not open XLSX: {e}")

        diag["sheets"] = wb.sheetnames

        # ── Locate expected sheet ─────────────────────────────────────────
        if expected_sheet not in wb.sheetnames:
            return err("err_submission_structurally_invalid",
                       f"Required sheet '{expected_sheet}' not found. "
                       f"Found: {wb.sheetnames}")
        ws = wb[expected_sheet]

        # ── Single pass: first row is the header, the rest are data ───────
        # One iterator avoids re-parsing the sheet twice (and the edge cases
        # that come with it).
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            return err("err_submission_structurally_invalid", "Header row is empty")

        # Strip the " *" required marker so identifiers match the canonical
        # model's field_name — fixes both the missing-headers check and the
        # record keys consumed by validate_upload (which look up by field_name).
        headers = [_undecorate(h) for h in header_row]
        missing_fields = [fn for fn in field_names if fn not in headers]
        if missing_fields:
            return err("err_submission_structurally_invalid",
                       f"Missing required headers: {missing_fields}")

        def _is_instruction_banner(row_tuple):
            # Treat the row as the builder's instruction banner iff every
            # non-empty cell equals the corresponding field's description.
            # Requires at least one real match, so an all-blank row isn't
            # mistaken for the banner (and an all-blank row would be dropped
            # as blank below regardless).
            if not has_descriptions:
                return False
            matched = False
            for i, cell in enumerate(row_tuple):
                if i >= len(headers) or not headers[i]:
                    continue
                cell_s = "" if cell is None else str(cell).strip()
                desc = descriptions.get(headers[i], "")
                if cell_s == "" and desc == "":
                    continue
                if cell_s != desc:
                    return False
                matched = True
            return matched

        # Supplier files are small (read_only is off anyway), so materialize the
        # remaining rows and drop the instruction banner if it's row 1 of data.
        data_rows = list(row_iter)
        banner_skipped = bool(data_rows) and _is_instruction_banner(data_rows[0])
        if banner_skipped:
            data_rows = data_rows[1:]
        diag["instruction_banner_skipped"] = banner_skipped

        rows = []
        for row_tuple in data_rows:
            record = {}
            for i, cell in enumerate(row_tuple):
                if i < len(headers) and headers[i]:
                    record[headers[i]] = cell
            # Keep only rows with at least one non-blank value.
            if any(v is not None and str(v).strip() != "" for v in record.values()):
                rows.append(record)

        wb.close()

        # default=str coerces datetime/date (openpyxl) so json.dumps won't raise;
        # the connector's check_data_type / check_data_format accept strings.
        rows_json = json.dumps(rows, default=str)
        diag["row_count"] = len(rows)
        return {
            "ok": True,
            "error_code": None,
            "error_message": None,
            "rows_json": rows_json,
            "row_count": len(rows),
            "log": json.dumps(diag),
        }

    except Exception as e:
        # Catch-all: the action ALWAYS returns the contract shape and never
        # leaks a raw traceback into the recipe. Full stack lands in `log`.
        diag["traceback"] = traceback.format_exc()
        return err("err_unexpected", f"Unhandled error during extraction: {e}")
