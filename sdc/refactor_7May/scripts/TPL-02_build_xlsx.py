"""
TPL-02 Build XLSX template

Entry point:
    main(input) -> dict

Expected `input` keys (per the recipe's input mapping):
    canonical_model_json : str   — JSON content of the canonical model (output of validate-config, read from FileStorage in recipe)
    variant_id           : str   — empty string for the base case
    customer_name        : str
    variant_name         : str   — resolved name; "base" when no variant_id
    protection_password  : str   — sheet/workbook password, supplied from the single source (NOT a module constant)

Returns:
    {
      'status'             : 'success' | 'empty_variant',
      'file_content'       : base64 str or None,
      'suggested_filename' : str or None,
      'metadata'           : dict or None,
      'error'              : {'code': str, 'message': str} or None,
    }

Failure modes:
    - 'empty_variant'         — soft outcome. Returned as a status value, with error.code = 'empty_variant'.
    - Any other BuildError    — raised as exception. These indicate Validate config let through a configuration it shouldn't have.

Sheet geometry (Data):
    row 1            = header (field names; required fields suffixed " *")
    row 2            = locked instruction banner (per-field description + any advisory hint)
    rows 3..N        = supplier data-entry rows
    All DV ranges and the cascade INDIRECT anchor are derived from DATA_START_ROW so they cannot drift apart.

Structure (two phases, in execution order):
    PLAN  (decide)  — interpret the canonical model into explicit records (ReferenceLayout, [FieldPlan]) with NO openpyxl writes.
                      All branching on data_type / data_format / lookup_name lives here, so field logic has a single home and is
                      testable without building a workbook.
    RENDER (emit)   — dumb writers that walk the records and emit the workbook, making no decisions.
    main collapses to: parse -> plan -> render -> serialize.
"""

import base64
import io
import json
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.protection import hash_password
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.datavalidation import DataValidation


# --- MODULE CONSTANTS + STYLES ----------------------------------------------------
DATA_ROWS               = 10000
HEADER_ROW              = 1
INSTRUCTION_ROW         = 2
DATA_START_ROW          = 3
REFERENCE_SHEET_NAME    = "Reference"
NAMED_RANGE_PREFIX      = "LU_"  # prefix every named range, avoids collisions with cell-reference patterns like "Q4"

_HEADER_FONT        = Font(bold=True, color="FFFFFF")
_HEADER_FILL        = PatternFill("solid", fgColor="4F81BD")
_HEADER_ALIGN       = Alignment(horizontal="left", vertical="center", wrap_text=True)
_REQUIRED_FILL      = PatternFill("solid", fgColor="C0504D")

_INSTRUCTION_FONT   = Font(italic=True, color="555555", size=8)
_INSTRUCTION_FILL   = PatternFill("solid", fgColor="F2F2F2")
_INSTRUCTION_ALIGN  = Alignment(horizontal="left", vertical="top", wrap_text=True)

_INSTRUCTION_BORDER = Border(
    left=Side(style="thin", color="555555"),
    right=Side(style="thin", color="555555"),
    top=Side(style="thin", color="555555"),
    bottom=Side(style="thin", color="555555"),
)
_HEADER_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="thin", color="555555"),
)

# Display-format patterns (used by _display_number_format)
_DEFAULT_DATE_FORMAT = "yyyy-mm-dd"
_DATE_MASK  = re.compile(r"\(([^)]+)\)")
_FLOAT_PREC = re.compile(r"float\s*\((\d+)\)")


# --- ERROR TYPE -------------------------------------------------------------------
# Caught only for empty_variant; everything else propagates
class BuildError(Exception):
    def __init__(self, code, message):
        self.code = code
        self.message = message
        super().__init__("[{0}] {1}".format(code, message))


# --- SHARED HELPERS ---------------------------------------------------------------
def _is_truthy(val):
    """Handles bool, int 1, "1", "TRUE", "true" (the canonical model may
    serialise booleans as numerics or strings)."""
    if val is True:
        return True
    if val == 1:
        return True
    s = str(val).strip().upper() if val is not None else ""
    return s in ("1", "TRUE", "YES")

def _safe_identifier(name):
    """Sanitize an arbitrary string into a safe identifier suffix."""
    return re.sub(r"[^A-Za-z0-9_]", "_", str(name))

def _format_label(field):
    """Normalized data_format label (lowercased, trimmed); '' when absent.
    The single source of truth for interpreting the analyst's shape dropdown."""
    return str(field.get("data_format") or "").strip().lower()


# --- SANITIZATION INVARIANT -------------------------------------------------------
"""
_sanitize_for_named_range (Python) and _build_indirect_formual (Excel sub) MUST apply the
same (char -> replacement) map, both derived from _compute_substitutions over the SAME parent
values. If drift occurs, dependent dropdowns will resolve to the wrong named range and silently
show the wrong options.

Changes to one must be mirrored in the other.
"""
_NAMED_RANGE_SAFE = re.compile(r"[A-Za-z0-9_]")

def _compute_substitutions(parent_values):
    disallowed = set()
    for value in parent_values:
        for char in str(value):
            if char == "_" or not _NAMED_RANGE_SAFE.match(char):
                disallowed.add(char)
    pairs = []
    if "_" in disallowed:
        pairs.append(("_", "_5F"))
    pairs.extend((c, "_{0:02X}".format(ord(c))) for c in sorted(c for c in disallowed if c != "_"))
    return pairs

def _sanitize_for_named_range(value, substitutions):
    """Python-side: apply the substitution map to a parent value."""
    result = str(value)
    for char, replacement in substitutions:
        result = result.replace(char, replacement)
    return result

def _build_indirect_formula(parent_cell_ref, substitutions, lookup_name):
    """Formula-side: build the SUBSTITUTE chain mirroring _sanitize_for_named_range."""
    expr = parent_cell_ref
    for char, replacement in substitutions:
        esc_char = char.replace('"', '""')
        esc_repl = replacement.replace('"', '""')
        expr = 'SUBSTITUTE({0},"{1}","{2}")'.format(expr, esc_char, esc_repl)

    prefix = "{0}{1}__".format(NAMED_RANGE_PREFIX, _safe_identifier(lookup_name))
    return 'INDIRECT("{0}"&{1})'.format(prefix, expr)

def _named_range_name(lookup_name, parent_value, substitutions):
    return "{0}{1}__{2}".format(
        NAMED_RANGE_PREFIX,
        _safe_identifier(lookup_name),
        _sanitize_for_named_range(parent_value, substitutions),
    )


# --- RECORD TYPES -----------------------------------------------------------------
# The contract between PLAN and RENDER
@dataclass
class ReferenceLayout:
    """Placement + named ranges for the (veryHidden) Reference sheet."""
    flat_columns: Dict[str, int]                       # lookup_name -> col_idx
    dependent_columns: Dict[Tuple[str, str], int]      # (lookup_name, parent_value) -> col_idx
    substitutions: List[Tuple[str, str]]               # the shared (char, replacement) map
    defined_ranges: List[Tuple[str, str]]              # (range_name, range_ref)


@dataclass
class FieldPlan:
    """Everything the Data sheet needs for ONE column, decided once."""
    col_idx: int
    col_letter: str
    field_id: Optional[str]
    header_text: str                                   # display + " *" if required
    is_required: bool
    is_locked: bool
    is_hidden: bool
    instruction_text: str                              # description + advisory hint
    number_format: Optional[str] = None                # display mask, or None
    data_validation: Optional[DataValidation] = None   # built in PLAN, attached in RENDER


# --- PLAN PHASE -------------------------------------------------------------------
# Decide (interpret the model into records; no openpyxl writes)
def _require_sheet_name(model):
    """Single source of truth for the data tab name. Reads from the canonical model's _meta."""
    name = (model.get("_meta") or {}).get("expected_sheet_name")
    if not name or not str(name).strip():
        raise BuildError(
            "MISSING_SHEET_NAME",
            "canonical model _meta.expected_sheet_name is missing or blank."
        )
    return str(name).strip()

def _resolve_fields(model, variant_id):
    """Return the variant's fields, sorted by position. Base case (no variant_id)
    returns all fields from the version."""
    all_fields = model.get("cfg_fields", [])
    if variant_id:
        included_ids = {
            vf["field_id"] for vf in model.get("cfg_variant_fields", [])
            if vf.get("variant_id") == variant_id
        }
        fields = [f for f in all_fields if f.get("field_id") in included_ids]
    else:
        fields = list(all_fields)
    return sorted(fields, key=lambda f: int(f.get("position", 0)))

def _index_lookups(cfg_lookups_rows, needed_names):
    """
    Convert the flat lookup-row list into a per-lookup structure.

    For each needed lookup, returns either:
        {'kind': 'flat', 'values': [v1, v2, ...]}
    or
        {'kind': 'dependent', 'groups': {parent: [child1, child2], ...}}

    Raises BuildError if a needed lookup has zero values.
    """
    flat = {}
    dependent = {}

    for row in cfg_lookups_rows:
        name = row.get("lookup_name")
        if name not in needed_names:
            continue
        value = row.get("valid_value")
        parent = row.get("parent_value")
        if parent:
            dependent.setdefault(name, {}).setdefault(parent, []).append(value)
        else:
            flat.setdefault(name, []).append(value)

    indexed = {}
    for name in needed_names:
        if name in dependent:
            indexed[name] = {"kind": "dependent", "groups": dependent[name]}
        elif name in flat:
            indexed[name] = {"kind": "flat", "values": flat[name]}
        else:
            raise BuildError(
                "LOOKUP_HAS_NO_VALUES",
                "Field references lookup '{0}' but it has no values.".format(name),
            )
    return indexed

def plan_reference(lookups) -> ReferenceLayout:
    """Decide column placement and named ranges for the Reference sheet. Builds
    the substitution map from the union of all dependent parent values. Pure;
    returns a ReferenceLayout. (Formerly _lay_out_reference, which returned a dict.)"""
    flat_columns = {}              # lookup_name -> col_idx
    dependent_columns = {}         # (lookup_name, parent_value) -> col_idx
    defined_ranges = []            # list of (range_name, range_ref)

    all_parents = []
    for lookup in lookups.values():
        if lookup["kind"] == "dependent":
            all_parents.extend(lookup["groups"].keys())
    substitutions = _compute_substitutions(all_parents)

    next_col = 1
    for name, lookup in lookups.items():
        if lookup["kind"] == "flat":
            flat_columns[name] = next_col
            col_letter = get_column_letter(next_col)
            range_ref = "'{0}'!${1}$2:${1}${2}".format(
                REFERENCE_SHEET_NAME, col_letter, 1 + len(lookup["values"])
            )
            defined_ranges.append(
                ("{0}FLAT_{1}".format(NAMED_RANGE_PREFIX, _safe_identifier(name)), range_ref)
            )
            next_col += 1
        else:
            for parent_value, children in lookup["groups"].items():
                dependent_columns[(name, parent_value)] = next_col
                col_letter = get_column_letter(next_col)
                range_ref = "'{0}'!${1}$2:${1}${2}".format(
                    REFERENCE_SHEET_NAME, col_letter, 1 + len(children)
                )
                defined_ranges.append(
                    (_named_range_name(name, parent_value, substitutions), range_ref)
                )
                next_col += 1

    return ReferenceLayout(
        flat_columns=flat_columns,
        dependent_columns=dependent_columns,
        substitutions=substitutions,
        defined_ranges=defined_ranges,
    )

def _display_number_format(field):
    """
    DISPLAY (number) format for a column, from the single data_format label.
    Display only: governs how Excel RENDERS a value — never what is accepted
    (that is the DV) nor what the server ingests (the stored value).
    Returns an Excel format code or None.
    """
    label     = _format_label(field)
    data_type = str(field.get("data_type") or "").strip().lower()

    if label.startswith("date") or data_type == "date":
        m = _DATE_MASK.search(field.get("data_format") or "")
        return m.group(1).strip().lower() if m else _DEFAULT_DATE_FORMAT

    if label == "percentage":
        m = _FLOAT_PREC.search(data_type)            # 'float (2)' -> 2
        decimals = int(m.group(1)) if m else 2
        return "0%" if decimals <= 0 else "0.{0}%".format("0" * decimals)

    if label == "currency":
        return "#,##0.00"   # symbol intentionally omitted — locale/currency unknown;
                            # cosmetic only, server is the authority for currency shape

    return None

def _advisory_hint(field):
    """
    Optional human-readable guidance appended to the locked instruction banner
    for constraints Excel cannot block in-cell. Keep hints generic and
    parse-free; the analyst's description stays the primary guidance lever, and
    VAL-01 stays the authority. Extend by adding cases.
    """
    label = _format_label(field)
    if label == "email address":
        return "Must be a valid email address (e.g. name@example.com)."
    # Future: a field carrying field_input_validation (regex) has no safe generic
    # hint — the raw pattern is meaningless to a supplier. Rely on the analyst's
    # description to state the expected format in words.
    return None

def _instruction_text(field):
    """Instruction-banner text: analyst description, plus any advisory hint."""
    parts = []
    desc = field.get("description")
    if desc and str(desc).strip():
        parts.append(str(desc).strip())
    hint = _advisory_hint(field)
    if hint:
        parts.append(hint)
    return "\n".join(parts)

def _build_advisory_validation(field, cell_range, display):
    """
    Strongest *advisory* in-cell check Excel can express for a non-lookup field.
    Returns a DataValidation or None.

    The authority for every field constraint is the server (VAL-01 validate_upload). This layer exists only to block easy mistakes at type-time
    and cut reject/resubmit round-trips. It deliberately does NOT attempt email / regex / uniqueness / cross-field shapes (Excel cannot block those
    in-cell), so they are surfaced as instruction hints (see _advisory_hint) and enforced server-side.

    Numeric-ness is driven by the *shape* as well as the primitive type, because 'currency' carries data_type 'string' in the model yet is numeric in practice
    (mirrors VAL-01's check_data_format).

    Range/length bounds are intentionally NOT parsed here. When the PRV chain emits resolved structured bounds at config-freeze, tighten this in ONE place
    so template and server never drift.
    """
    label     = _format_label(field)
    data_type = str(field.get("data_type") or "").strip().lower()
    allow_blank = not _is_truthy(field.get("required"))
    top_left  = cell_range.split(":")[0]

    is_numeric = data_type in ("integer", "float (2)") or label in ("currency", "percentage")
    is_integer = data_type == "integer"
    is_date    = data_type == "date" or label.startswith("date")

    # Numeric (incl. currency / percentage)
    if is_numeric:
        if is_integer:
            formula = "AND(ISNUMBER({0}), {0}=INT({0}))".format(top_left)
        else:
            formula = "ISNUMBER({0})".format(top_left)
        dv = DataValidation(type="custom", formula1=formula, allow_blank=allow_blank)
        dv.error = "{0} must be a number.".format(display)
        dv.errorTitle = "Invalid value"
        dv.showErrorMessage = True
        dv.add(cell_range)
        return dv

    # Date
    if is_date:
        dv = DataValidation(
            type="date", operator="greaterThanOrEqual",
            formula1="1900-01-01", allow_blank=allow_blank,
        )
        dv.error = "{0} must be a valid date.".format(display)
        dv.errorTitle = "Invalid date"
        dv.showErrorMessage = True
        dv.add(cell_range)
        return dv

    # string / email / boolean / unshaped
    # No in-cell block. Guidance (if any) goes to the instruction banner.
    return None

def _plan_validation(field, cell_range, field_col_index, lookups, layout):
    """
    Decide the single DataValidation for one field (or None). The three branches
    (dependent dropdown (INDIRECT), flat dropdown (named range), and type/format
    (advisory)) are the former body of _apply_data_validations, unchanged in
    logic; they now RETURN the DV rather than attaching it, so the decision is
    testable without a worksheet. This is the only place field validation is decided.
    """
    field_id    = field.get("field_id")
    display     = field.get("field_name", field_id or "")
    lookup_name = field.get("lookup_name")
    parent_id   = field.get("cascade_parent_field_id")

    # Dependent dropdown
    if lookup_name and parent_id:
        lookup = lookups.get(lookup_name)
        if lookup is None or lookup["kind"] != "dependent":
            raise BuildError(
                "DEPENDENT_FIELD_FLAT_LOOKUP",
                "Field '{0}' declares parent '{1}' but lookup '{2}' is flat or missing.".format(
                    field_id, parent_id, lookup_name
                ),
            )
        parent_col_idx = field_col_index.get(parent_id)
        if parent_col_idx is None:
            raise BuildError(
                "PARENT_FIELD_NOT_IN_VARIANT",
                "Dependent field '{0}' references parent '{1}' which is not in the resolved field set.".format(
                    field_id, parent_id
                ),
            )
        parent_letter = get_column_letter(parent_col_idx)
        # Anchor the INDIRECT at the first DATA row of the parent column. Must match
        # cell_range's start row (DATA_START_ROW); openpyxl resolves the relative
        # reference from the range's top-left, so the anchor and the range start are
        # a coupled pair — both via the constant.
        formula1 = _build_indirect_formula(
            "{0}{1}".format(parent_letter, DATA_START_ROW),
            layout.substitutions,
            lookup_name,
        )
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        dv.error = "Select a valid {0}".format(display)
        dv.errorTitle = "Invalid selection"
        dv.showErrorMessage = True
        dv.add(cell_range)
        return dv

    # Flat dropdown
    if lookup_name:
        range_name = "{0}FLAT_{1}".format(NAMED_RANGE_PREFIX, _safe_identifier(lookup_name))
        dv = DataValidation(type="list", formula1="={0}".format(range_name), allow_blank=True)
        dv.error = "Select a valid {0}".format(display)
        dv.errorTitle = "Invalid selection"
        dv.showErrorMessage = True
        dv.add(cell_range)
        return dv

    # Type / format constraint (advisory)
    return _build_advisory_validation(field, cell_range, display)

def plan_fields(fields, lookups, layout) -> List[FieldPlan]:
    """
    Interpret each field into a FieldPlan. Two passes so a dependent dropdown can
    reference its parent column without a mid-loop index rebuild:
      pass 1 — assign columns + static attributes (and the field_id -> col map)
      pass 2 — build each field's data validation (may reference a parent column)
    No openpyxl writes happen here; this is pure decision-making.
    """
    last_row = DATA_START_ROW + DATA_ROWS - 1

    # Pass 1: columns + static attributes
    plans: List[FieldPlan] = []
    for col_idx, f in enumerate(fields, start=1):
        display  = f.get("field_name", f.get("field_id", ""))
        required = _is_truthy(f.get("required"))
        plans.append(FieldPlan(
            col_idx=col_idx,
            col_letter=get_column_letter(col_idx),
            field_id=f.get("field_id"),
            header_text=display + (" *" if required else ""),
            is_required=required,
            is_locked=_is_truthy(f.get("read_only")),
            is_hidden=_is_truthy(f.get("hidden")),
            instruction_text=_instruction_text(f),
            number_format=_display_number_format(f),
        ))

    field_col_index = {p.field_id: p.col_idx for p in plans}

    # Pass 2: validations (may reference a parent column)
    for plan, f in zip(plans, fields):
        cell_range = "{0}{1}:{0}{2}".format(plan.col_letter, DATA_START_ROW, last_row)
        plan.data_validation = _plan_validation(f, cell_range, field_col_index, lookups, layout)

    return plans

# --- RENDER PHASE -----------------------------------------------------------------
# Emit, walk the records, and write the workbook. No decisions are made here.
def _create_workbook(sheet_name):
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = sheet_name
    ref_ws = wb.create_sheet(REFERENCE_SHEET_NAME)
    ref_ws.sheet_state = "veryHidden"  # "hidden" can be unhidden from the Excel UI, "veryHidden" cannot
    return wb, data_ws, ref_ws

def _format_header_cell(cell, required):
    cell.font = _HEADER_FONT
    cell.fill = _REQUIRED_FILL if required else _HEADER_FILL
    cell.alignment = _HEADER_ALIGN
    cell.border = _HEADER_BORDER

def render_data_sheet(data_ws, plans):
    """
    Emit the Data sheet from FieldPlans. Pure rendering — every decision (header
    text, lock state, instruction text, number format, validation) was already
    made in plan_fields. This one walk replaces the former four loops
    (_write_data_header, _write_instruction_row, _apply_column_formats, and the
    DV-attach pass of _apply_data_validations).
    """
    for p in plans:
        # Header (row 1)
        hc = data_ws.cell(row=HEADER_ROW, column=p.col_idx, value=p.header_text)
        _format_header_cell(hc, p.is_required)
        data_ws.column_dimensions[p.col_letter].width = max(len(p.header_text) + 15, 14)

        # Hide hidden columns
        if p.is_hidden:
            data_ws.column_dimensions[p.col_letter].hidden = True

        # Instruction banner (row 2) — explicitly locked, so blank cells in
        # otherwise-editable columns stay non-editable under sheet protection.
        ic = data_ws.cell(row=INSTRUCTION_ROW, column=p.col_idx, value=p.instruction_text)
        ic.font = _INSTRUCTION_FONT
        ic.fill = _INSTRUCTION_FILL
        ic.border = _INSTRUCTION_BORDER
        ic.alignment = _INSTRUCTION_ALIGN
        ic.protection = Protection(locked=True)

        # Display format (column default; unmaterialised entry cells inherit it)
        if p.number_format:
            data_ws.column_dimensions[p.col_letter].number_format = p.number_format

        # Validation (built during PLAN; only attached here)
        if p.data_validation is not None:
            data_ws.add_data_validation(p.data_validation)

    data_ws.row_dimensions[INSTRUCTION_ROW].height = 56
    # Pin both the header and the instruction banner.
    data_ws.freeze_panes = "A{0}".format(DATA_START_ROW)

def render_reference_sheet(ref_ws, lookups, layout):
    """Emit the (veryHidden) Reference sheet: one column per flat lookup and per
    dependent (lookup, parent) group, headed by the sanitized parent token so the
    named ranges resolve. (Formerly _write_reference_content.)"""
    substitutions = layout.substitutions

    for name, lookup in lookups.items():
        if lookup["kind"] == "flat":
            col_idx = layout.flat_columns[name]
            ref_ws.cell(row=1, column=col_idx, value=name).font = _HEADER_FONT
            for row_offset, value in enumerate(lookup["values"], start=2):
                ref_ws.cell(row=row_offset, column=col_idx, value=value)
        else:
            for parent_value, children in lookup["groups"].items():
                col_idx = layout.dependent_columns[(name, parent_value)]
                sanitized = _sanitize_for_named_range(parent_value, substitutions)
                ref_ws.cell(row=1, column=col_idx, value=sanitized).font = _HEADER_FONT
                for row_offset, child_value in enumerate(children, start=2):
                    ref_ws.cell(row=row_offset, column=col_idx, value=child_value)

def register_defined_names(wb, layout):
    """Attach every (name, range_ref) from the layout as a workbook defined name.
    (Formerly _register_defined_names.)"""
    for name, range_ref in layout.defined_ranges:
        wb.defined_names[name] = DefinedName(name=name, attr_text=range_ref)

def apply_protection(wb, data_ws, ref_ws, plans, password):
    """
    Lock the template so suppliers edit only intended data-entry cells. Each field column's default style
    is set 'unlocked' unless the plan is locked; empty body cells inherit the column default, so editable
    columns are writable without materialising 1k cells/field. Header cells carry their own format and retain
    the default locked=True. Instruction-banner cells are explicitly locked in render_data_sheet, so they stay
    non-editable even in unlocked columns.
    Sheet and workbook protection are password-set so a supplier can't strip them under Review > Unprotect Sheet.

    `password` is supplied by the caller (single source of truth), NOT a module constant, so build (TPL-02)
    and re-seal (INC-02) cannot drift apart. The attribute-assignment path below was verified to serialize the
    correct legacy 16-bit hash (identical to SheetProtection(...) and set_password()), so it is left as-is.
    """
    # Tripwire: a blank/missing password yields protection no one can lift later
    # (a silent lockout). Fail loud at build instead of shipping a dead-locked file.
    if not password or not str(password).strip():
        raise BuildError(
            "MISSING_PROTECTION_PASSWORD",
            "Sheet/workbook password was empty; refusing to build a template that "
            "cannot be unprotected. Check the password source."
        )
    password = str(password)  # exact value; do NOT strip — trailing chars are caller intent

    for p in plans:
        if p.is_locked or p.is_hidden:
            continue  # locked OR hidden col inherits default locked=True
        data_ws.column_dimensions[p.col_letter].protection = Protection(locked=False)

    data_ws.protection.sheet = True
    data_ws.protection.password = password
    ref_ws.protection.sheet = True
    ref_ws.protection.password = password

    wb.security = WorkbookProtection(lockStructure=True)
    wb.security.set_workbook_password(password, already_hashed=False)


# --- OUTPUT HELPERS----------------------------------------------------------------
def _serialize_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _build_filename(customer_name, variant_name):
    safe_customer = _safe_identifier(customer_name)
    safe_variant = _safe_identifier(variant_name) if variant_name else "base"
    stamp = datetime.utcnow().strftime("%Y%m%d")
    return "{0}_{1}_{2}.xlsx".format(safe_customer, safe_variant, stamp)

def _empty_variant_outcome():
    return {
        "status": "empty_variant",
        "file_content": None,
        "suggested_filename": None,
        "metadata": None,
        "error": {
            "code": "empty_variant",
            "message": (
                "Variant resolved to zero fields. Check the variant's "
                "field associations or the version's field definitions."
            ),
        },
    }


# --- MAIN -------------------------------------------------------------------------
# Parse, plan, render, then serialize.
def main(input):
    # Parse the canonical model. Malformed JSON here is a precondition violation that should propagate.
    model = json.loads(input["canonical_model_json"])
    sheet_name = _require_sheet_name(model)

    variant_id = input.get("variant_id") or None
    customer_name = input["customer_name"]
    variant_name = input.get("variant_name") or "base"
    protection_password = input.get("protection_password")   # single source of truth

    # 1. PLAN (decide): interpret config into records; no openpyxl writes
    fields = _resolve_fields(model, variant_id)
    if not fields:
        return _empty_variant_outcome()

    needed_lookup_names = {f["lookup_name"] for f in fields if f.get("lookup_name")}
    lookups = _index_lookups(model.get("cfg_lookups", []), needed_lookup_names)
    layout = plan_reference(lookups)
    plans = plan_fields(fields, lookups, layout)

    # 2. RENDER (emit): build the workbook from the records; no decisions
    wb, data_ws, ref_ws = _create_workbook(sheet_name)
    render_data_sheet(data_ws, plans)
    render_reference_sheet(ref_ws, lookups, layout)
    register_defined_names(wb, layout)
    apply_protection(wb, data_ws, ref_ws, plans, protection_password)

    # 3. SERIALIZE & COMPOSE OUTPUT
    file_bytes = _serialize_to_bytes(wb)
    return {
        "status": "success",
        "file_content": base64.b64encode(file_bytes).decode("ascii"),
        "suggested_filename": _build_filename(customer_name, variant_name),
        "metadata": {
            "sheet_names": [sheet_name, REFERENCE_SHEET_NAME],
            "byte_size": len(file_bytes),
            "row_count": 0,
            "field_count": len(plans),
            "locked_field_count": sum(1 for p in plans if p.is_locked),
            "hidden_field_count": sum(1 for p in plans if p.is_hidden),
        },
        "error": None,
        # --- TEMPORARY DIAGNOSTIC (remove after confirming the password) ---
        "debug_pw_repr": repr(protection_password),
        "debug_pw_hash": hash_password(str(protection_password or "")),
    }
