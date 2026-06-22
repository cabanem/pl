"""
Hydrate a Supplier Data Collection template with one supplier's slice of seed data.

Opens the EXISTING blank template (the TPL-02 output, preserving dropdowns, named ranges, the veryHidden Reference sheet, 
and per-column sheet protection), filters the seed CSV down to a single supplier, and writes the matching rows into the
"Data Entry" sheet by matching seed column names to the template's row-1 headers.

Aligned to TPL-02 geometry:
    row 1      = field-name header (TPL-02 appends ' *' to required fields)
    row 2      = locked instruction banner (NOT matched, never written)
    rows 3..N  = supplier data-entry rows

Template fields with no seed column are left blank; seed columns with no template field are ignored. `hydrate_template` is Workato-agnostic 
and unit-testable; the entrypoint at the bottom is a thin adapter.
"""

import base64
import binascii
import csv
import io
import re

import openpyxl
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter


# --- Template geometry (must track TPL-02) ----------------------------------
DATA_ENTRY_SHEET 	= "Data"
HEADER_ROW 			= 1           	# field names live here (and only here)
DATA_START_ROW 		= 3       		# first data row, immediately below the instruction banner

# TPL-02 marks required fields by appending ' *' to the header text. Strip it so 'supplier_name' (seed) matches 
# 'supplier_name *' (template). Applied to BOTH sides, so it also matches if an analyst copied the asterisks into the seed file.
_REQUIRED_MARKER = re.compile(r"\s*\*\s*$")


def _field_key(name):
    """Canonical key for matching a header/column name across seed and template."""
    return _REQUIRED_MARKER.sub("", str(name if name is not None else "")).strip()


# --- File-boundary helpers (the bytes / base64 handling we settled on) ------
def _to_bytes(content):
    """Normalize a Workato file datapill to bytes without corrupting binary."""
    if isinstance(content, (bytes, bytearray)):
        return bytes(content)
    if isinstance(content, str):
        return content.encode("latin-1")
    raise TypeError(f"Unexpected file content type: {type(content)}")


def _decode_xlsx(content):
    """Return raw XLSX (zip) bytes. Branch on the magic number instead of
    blindly base64-decoding, which would mangle already-raw bytes."""
    raw = _to_bytes(content)
    magic = binascii.hexlify(raw[:4]).decode("ascii")
    if magic == "504b0304":      # 'PK\x03\x04' -> raw zip / xlsx, use as-is
        return raw
    if magic == "55457344":      # 'UEsD'       -> base64-encoded xlsx, decode
        return base64.b64decode(raw)
    raise ValueError(f"Template is not a recognizable XLSX (magic={magic})")


def _decode_csv_text(content):
    """Seed CSV arrives as the file's raw bytes (UTF-8, possibly with a BOM)."""
    return _to_bytes(content).decode("utf-8-sig")


def _column_locked(ws, col_letter, default=False):
    """Read a column's locked state from the loaded template so seeded cells can
    inherit it. Defaults to editable (False): seed pre-fills editable fields, and
    VAL-01 remains the authority on read-only fields regardless of cell locking."""
    try:
        dim = ws.column_dimensions.get(col_letter)
        if dim is not None and dim.protection is not None and dim.protection.locked is not None:
            return bool(dim.protection.locked)
    except Exception:
        pass
    return default


def _read_seed_rows(content):
    """Return (rows, fieldnames) from the seed file, dispatching on magic number.
    rows: list of dicts keyed by the ORIGINAL header text (CSV-DictReader shape),
    so downstream matching by key_to_header[...] is identical for both formats."""
    raw = _to_bytes(content)
    if raw[:4] == b"PK\x03\x04":            # XLSX / zip
        return _read_seed_rows_xlsx(raw)
    if raw[:2] == b"PK":
        raise ValueError("seed file looks like a zip but not a standard XLSX")
    return _read_seed_rows_csv(raw)          # text CSV


def _read_seed_rows_csv(raw):
    reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
    fieldnames = list(reader.fieldnames or [])
    return list(reader), fieldnames


def _read_seed_rows_xlsx(raw):
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None) or ()
        fieldnames = [str(h) if h is not None else "" for h in header]
        rows = []
        for r in it:
            row = {}
            for j, h in enumerate(fieldnames):
                if h == "":
                    continue
                v = r[j] if j < len(r) else None
                row[h] = _cell_text(v)       # match CSV's all-text shape
            rows.append(row)
        return rows, fieldnames
    finally:
        wb.close()


def _cell_text(value):
    """Stringify openpyxl cells the way csv yields text. Collapses 1234.0 -> '1234'
    so a numeric index/match value compares equal to its CSV form."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
  
# --- Pure core --------------------------------------------------------------
def hydrate_template(template_bytes, seed_csv_bytes, index_key, match_value):
    """Return (xlsx_bytes, diagnostics)."""

	# 1) Parse the seed file (CSV or XLSX). Map canonical key -> original header.
    seed_rows, seed_fieldnames = _read_seed_rows(seed_csv_bytes)
    key_to_header = {}
    for header in seed_fieldnames:
        key_to_header[_field_key(header)] = header

    index_norm = _field_key(index_key)
    if index_norm not in key_to_header:
        raise ValueError(
            f"index_key '{index_key}' not in seed columns: {list(key_to_header)}"
        )
    index_header = key_to_header[index_norm]

    target = (match_value or "").strip()
    matched = [
        row for row in seed_rows
        if str(row.get(index_header) or "").strip() == target
    ]

    # 2) Open the template, preserving every structural feature.
    wb = openpyxl.load_workbook(io.BytesIO(_decode_xlsx(template_bytes)))
    if DATA_ENTRY_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{DATA_ENTRY_SHEET}' not found; sheets: {wb.sheetnames}"
        )
    ws = wb[DATA_ENTRY_SHEET]

    # 3) Map seed key -> (column, letter, locked) by reading ONLY the row-1 header.
    col_for_key = {}
    for col in range(1, ws.max_column + 1):
        key = _field_key(ws.cell(row=HEADER_ROW, column=col).value)
        if key and key in key_to_header and key not in col_for_key:
            letter = get_column_letter(col)
            col_for_key[key] = (col, letter, _column_locked(ws, letter))

    writable = list(col_for_key.keys())                       # template column order
    skipped = [k for k in key_to_header if k not in col_for_key and k != index_norm]

    # 4) Write matched rows. Skip empties (blank cells stay truly empty). Mirror each column's locked state 
    # (so seeded cells honor the template's editable / read-only intent under sheet protection).
    for i, row in enumerate(matched):
        excel_row = DATA_START_ROW + i
        for key in writable:
            col, letter, locked = col_for_key[key]
            value = row.get(key_to_header[key], "")
            if value is None or value == "":
                continue
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.protection = Protection(locked=locked)

    # 5) Serialize back to XLSX bytes.
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    diagnostics = {
        "match_value": target,
        "matched_row_count": len(matched),
        "fields_written": writable,
        "seed_fields_ignored": skipped,   # in seed, absent from template
        "first_data_row": DATA_START_ROW,
        "last_data_row": (DATA_START_ROW + len(matched) - 1) if matched else None,
    }
    return xlsx_bytes, diagnostics


# --- Workato entrypoint (thin adapter) --------------------------------------
# Inputs:
#   template_file : get_file_contents(template_path)
#   seed_file     : get_file_contents(seed_data_file_path)
#   index_key     : parameters.seed_data.index_key
#   match_value   : parameters.supplier.supplier_name
#
# Output seeded_template_base64 survives any Workato boundary; in step 7 set the
# upload's file_content to:  <pill>.decode_base64
def main(input):
    xlsx_bytes, diag = hydrate_template(
        template_bytes=input["template_file"],
        seed_csv_bytes=input["seed_file"],
        index_key=input["index_key"],
        match_value=input["match_value"],
    )
    return {
        "seeded_template_base64": base64.b64encode(xlsx_bytes).decode("ascii"),
        "matched_row_count": diag["matched_row_count"],
        "fields_written": ", ".join(diag["fields_written"]),
        "seed_fields_ignored": ", ".join(diag["seed_fields_ignored"]),
        "first_data_row": diag["first_data_row"],
        "last_data_row": diag["last_data_row"] if diag["last_data_row"] is not None else 0,
    }
