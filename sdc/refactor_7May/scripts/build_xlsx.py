"""
Build XLSX Template — production module.

Implements the Build XLSX Template capability described in
sdc-capability-build-xlsx-template-v1.md.

Entry point: build_xlsx_template(payload: dict) -> dict

Designed to drop into a Workato Python action. The entry point takes a
plain dict (so it composes with Workato's datapill mapping) and returns
a plain dict with the base64-encoded workbook bytes, a suggested
filename, and metadata.

The substage numbering in comments mirrors the deep dive doc.
"""

from __future__ import annotations

import base64
import io
import re
from dataclasses import dataclass, field as dataclass_field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


# ──────────────────────────────────────────────────────────────────────
# Module constants
# ──────────────────────────────────────────────────────────────────────

DATA_ROWS = 1000                    # number of pre-validated rows on the data entry sheet
DATA_SHEET_NAME = "Data Entry"
REFERENCE_SHEET_NAME = "Reference"
NAMED_RANGE_PREFIX = "LU_"          # prefix on every dependent-lookup named range,
                                    # to avoid collisions with cell-reference patterns
                                    # (e.g. a parent value of "Q4" would otherwise
                                    # collide with the Q4 cell reference)


# ──────────────────────────────────────────────────────────────────────
# Input data shapes
# ──────────────────────────────────────────────────────────────────────

@dataclass
class FieldDef:
    field_id: str
    display_name: str
    position: int
    data_type: str                                  # "text" | "number" | "date" | "boolean"
    required: bool = False
    lookup_name: Optional[str] = None               # set for dropdown fields
    parent_field_id: Optional[str] = None           # set for dependent-dropdown fields
    data_format: Dict[str, Any] = dataclass_field(default_factory=dict)
    # data_format keys (all optional, type-specific):
    #   text:    "max_length" -> int
    #   number:  "min" -> number, "max" -> number, "integer" -> bool
    #   date:    "min" -> "YYYY-MM-DD", "max" -> "YYYY-MM-DD"


@dataclass
class LookupDef:
    lookup_name: str
    # Exactly one of these is populated:
    flat_values: Optional[List[str]] = None
    parent_groups: Optional[Dict[str, List[str]]] = None
    # parent_groups: { parent_value -> [child_value, ...] }

    @property
    def is_dependent(self) -> bool:
        return self.parent_groups is not None


@dataclass
class BuildInput:
    fields: List[FieldDef]
    lookups: Dict[str, LookupDef]
    client_name: str
    variant_name: Optional[str] = None              # None == base case (no variant)


# ──────────────────────────────────────────────────────────────────────
# Output data shapes
# ──────────────────────────────────────────────────────────────────────

@dataclass
class BuildOutput:
    file_content_base64: str
    suggested_filename: str
    metadata: Dict[str, Any]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "file_content": self.file_content_base64,
            "suggested_filename": self.suggested_filename,
            "metadata": self.metadata,
        }


# ──────────────────────────────────────────────────────────────────────
# Shared sanitization — the load-bearing invariant
# ──────────────────────────────────────────────────────────────────────
#
# Named ranges in Excel allow letters, digits, underscores, and a few
# other characters but disallow spaces, &, /, (, ), -, and many others.
# When a dependent dropdown's parent value contains a disallowed char,
# the named range built from that value has to be sanitized.
#
# The catch: the INDIRECT formula in the data validation rule has to
# compute the SAME sanitized string from the cell value at runtime. If
# the Python sanitizer and the formula's SUBSTITUTE chain don't agree,
# the dropdown silently resolves to nothing.
#
# The fix: derive the substitution map ONCE from the actual parent
# values, then use it to (a) name the ranges in Python and (b) build
# the SUBSTITUTE chain in the formula. By construction they cannot drift.

_NAMED_RANGE_SAFE = re.compile(r"[A-Za-z0-9_]")


def compute_substitutions(parent_values: List[str]) -> List[Tuple[str, str]]:
    """
    Scan all parent values, find characters that aren't valid in named
    range names, return a sorted list of (char, '_') pairs.

    Sorted ordering is important: it makes the SUBSTITUTE chain
    deterministic, and stable across runs given the same input.
    """
    disallowed: set = set()
    for value in parent_values:
        for char in str(value):
            if not _NAMED_RANGE_SAFE.match(char):
                disallowed.add(char)
    return [(char, "_") for char in sorted(disallowed)]


def sanitize_for_named_range(value: str,
                             substitutions: List[Tuple[str, str]]) -> str:
    """
    Apply the substitution map to a parent value, producing the
    suffix used in the named range. Same logic the SUBSTITUTE chain
    will apply at runtime.
    """
    result = str(value)
    for char, replacement in substitutions:
        result = result.replace(char, replacement)
    return result


def build_substitute_formula(cell_ref: str,
                             substitutions: List[Tuple[str, str]]) -> str:
    """
    Build the SUBSTITUTE chain that mirrors `sanitize_for_named_range`
    at formula time. Wrapped so that INDIRECT can resolve the named
    range from the parent cell.

    Returns the formula1 string for a list-type DataValidation.
    """
    expr = cell_ref
    for char, replacement in substitutions:
        # Escape any literal " in the char (rare but possible)
        escaped_char = char.replace('"', '""')
        escaped_repl = replacement.replace('"', '""')
        expr = f'SUBSTITUTE({expr},"{escaped_char}","{escaped_repl}")'
    return f'INDIRECT("{NAMED_RANGE_PREFIX}"&{expr})'


def named_range_name(parent_value: str,
                     substitutions: List[Tuple[str, str]]) -> str:
    """Compose the full named range identifier for a parent value."""
    return f"{NAMED_RANGE_PREFIX}{sanitize_for_named_range(parent_value, substitutions)}"


# ──────────────────────────────────────────────────────────────────────
# Substage 1 — Resolve fields (and validate the structural assumptions
# this capability is allowed to assume)
# ──────────────────────────────────────────────────────────────────────

def _resolve_fields(build_input: BuildInput) -> List[FieldDef]:
    """
    The variant filter is the caller's responsibility per the deep
    dive — this capability is called with the already-filtered field
    list. We only sort by position here.
    """
    if not build_input.fields:
        raise BuildError("EMPTY_FIELD_LIST",
                         "No fields supplied to build template. "
                         "An empty variant should be caught upstream.")
    return sorted(build_input.fields, key=lambda f: f.position)


# ──────────────────────────────────────────────────────────────────────
# Substage 2 — Resolve lookups actually used by the resolved fields
# ──────────────────────────────────────────────────────────────────────

def _resolve_lookups(
        fields: List[FieldDef],
        lookups: Dict[str, LookupDef]) -> Dict[str, LookupDef]:
    """
    Return only the lookups referenced by at least one field. Other
    lookups in the configuration are ignored — they may belong to a
    different variant.
    """
    needed = {f.lookup_name for f in fields if f.lookup_name}
    resolved: Dict[str, LookupDef] = {}
    for name in needed:
        if name not in lookups:
            # Validate config should have caught this; we don't diagnose,
            # we fail loudly.
            raise BuildError("LOOKUP_NOT_FOUND",
                             f"Field references lookup '{name}' which is "
                             "not in the supplied configuration.")
        resolved[name] = lookups[name]
    return resolved


# ──────────────────────────────────────────────────────────────────────
# Substage 3 — Lay out the reference sheet
# ──────────────────────────────────────────────────────────────────────

@dataclass
class ReferenceLayout:
    # Per flat lookup: { lookup_name -> column index on reference sheet }
    flat_columns: Dict[str, int]
    # Per dependent lookup: { (lookup_name, parent_value) -> column index }
    dependent_columns: Dict[Tuple[str, str], int]
    # The substitution map used to sanitize parent values; reused by
    # the formula builder so the names and the formula stay in sync.
    substitutions: List[Tuple[str, str]]
    # Names of the defined ranges to register, with their cell ranges.
    defined_ranges: List[Tuple[str, str]]


def _lay_out_reference(
        lookups: Dict[str, LookupDef]) -> ReferenceLayout:
    """
    Decide which column each lookup (or parent-group) occupies on the
    reference sheet. Build the substitution map from the union of all
    parent values across all dependent lookups — derived once so the
    sanitizer and the formula builder share a single source of truth.
    """
    flat_columns: Dict[str, int] = {}
    dependent_columns: Dict[Tuple[str, str], int] = {}
    defined_ranges: List[Tuple[str, str]] = []

    # Collect all parent values across all dependent lookups.
    all_parents: List[str] = []
    for lookup in lookups.values():
        if lookup.is_dependent and lookup.parent_groups:
            all_parents.extend(lookup.parent_groups.keys())

    substitutions = compute_substitutions(all_parents)

    next_col = 1
    for lookup in lookups.values():
        if not lookup.is_dependent:
            # Flat: one column for the whole lookup.
            flat_columns[lookup.lookup_name] = next_col
            col_letter = get_column_letter(next_col)
            range_ref = (
                f"'{REFERENCE_SHEET_NAME}'!${col_letter}$2:"
                f"${col_letter}${1 + len(lookup.flat_values or [])}"
            )
            # Flat lookups get a named range too, for consistency.
            defined_ranges.append(
                (f"{NAMED_RANGE_PREFIX}FLAT_{_safe_identifier(lookup.lookup_name)}",
                 range_ref)
            )
            next_col += 1
        else:
            # Dependent: one column per parent value.
            for parent_value, children in (lookup.parent_groups or {}).items():
                dependent_columns[(lookup.lookup_name, parent_value)] = next_col
                col_letter = get_column_letter(next_col)
                range_ref = (
                    f"'{REFERENCE_SHEET_NAME}'!${col_letter}$2:"
                    f"${col_letter}${1 + len(children)}"
                )
                defined_ranges.append(
                    (named_range_name(parent_value, substitutions), range_ref)
                )
                next_col += 1

    return ReferenceLayout(
        flat_columns=flat_columns,
        dependent_columns=dependent_columns,
        substitutions=substitutions,
        defined_ranges=defined_ranges,
    )


def _safe_identifier(name: str) -> str:
    """Sanitize a lookup name for use as an identifier suffix."""
    return re.sub(r"[^A-Za-z0-9_]", "_", name)


# ──────────────────────────────────────────────────────────────────────
# Substages 4–6 — Workbook construction, header row, reference content
# ──────────────────────────────────────────────────────────────────────

# Styling constants — kept simple on purpose. Header cells are
# distinguished by font weight and a soft fill; required fields by an
# asterisk in the display name. No fancy theming.

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_REQUIRED_HEADER_FONT = Font(bold=True, color="FFFFFF")
_REQUIRED_HEADER_FILL = PatternFill("solid", fgColor="C0504D")


def _create_workbook() -> Tuple[Workbook, Worksheet, Worksheet]:
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = DATA_SHEET_NAME
    ref_ws = wb.create_sheet(REFERENCE_SHEET_NAME)
    ref_ws.sheet_state = "hidden"
    return wb, data_ws, ref_ws


def _write_data_header(data_ws: Worksheet,
                       fields: List[FieldDef]) -> None:
    for col_idx, field in enumerate(fields, start=1):
        header_text = field.display_name + (" *" if field.required else "")
        cell = data_ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = _REQUIRED_HEADER_FONT if field.required else _HEADER_FONT
        cell.fill = _REQUIRED_HEADER_FILL if field.required else _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        data_ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(header_text) + 4, 14
        )
    data_ws.freeze_panes = "A2"


def _write_reference_content(ref_ws: Worksheet,
                             lookups: Dict[str, LookupDef],
                             layout: ReferenceLayout,
                             substitutions: List[Tuple[str, str]]) -> None:
    # Write a small header row on the reference sheet too — purely for
    # human inspection if someone unhides it; not referenced by any
    # validation rule.
    for lookup in lookups.values():
        if not lookup.is_dependent:
            col_idx = layout.flat_columns[lookup.lookup_name]
            ref_ws.cell(row=1, column=col_idx, value=lookup.lookup_name).font = _HEADER_FONT
            for row_offset, value in enumerate(lookup.flat_values or [], start=2):
                ref_ws.cell(row=row_offset, column=col_idx, value=value)
        else:
            for parent_value, children in (lookup.parent_groups or {}).items():
                col_idx = layout.dependent_columns[(lookup.lookup_name, parent_value)]
                # The header is the sanitized parent name (matches the
                # named range suffix), so an unhider can see the link.
                sanitized = sanitize_for_named_range(parent_value, substitutions)
                ref_ws.cell(row=1, column=col_idx, value=sanitized).font = _HEADER_FONT
                for row_offset, child_value in enumerate(children, start=2):
                    ref_ws.cell(row=row_offset, column=col_idx, value=child_value)


def _register_defined_names(wb: Workbook,
                            layout: ReferenceLayout) -> None:
    for name, range_ref in layout.defined_ranges:
        defined = DefinedName(name=name, attr_text=range_ref)
        wb.defined_names[name] = defined


# ──────────────────────────────────────────────────────────────────────
# Substage 7 — Apply data validation rules
# ──────────────────────────────────────────────────────────────────────

def _apply_data_validations(data_ws: Worksheet,
                            fields: List[FieldDef],
                            lookups: Dict[str, LookupDef],
                            layout: ReferenceLayout) -> None:
    # field_id -> column index on the data sheet
    field_col_index: Dict[str, int] = {
        f.field_id: idx for idx, f in enumerate(fields, start=1)
    }

    last_row = 1 + DATA_ROWS  # row 1 is header

    for col_idx, field in enumerate(fields, start=1):
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{last_row}"

        # ── Dependent dropdown ──
        if field.lookup_name and field.parent_field_id:
            lookup = lookups[field.lookup_name]
            if not lookup.is_dependent:
                raise BuildError(
                    "DEPENDENT_FIELD_FLAT_LOOKUP",
                    f"Field '{field.field_id}' declares parent "
                    f"'{field.parent_field_id}' but lookup "
                    f"'{field.lookup_name}' is flat.",
                )

            parent_col_idx = field_col_index.get(field.parent_field_id)
            if parent_col_idx is None:
                raise BuildError(
                    "PARENT_FIELD_NOT_IN_VARIANT",
                    f"Dependent field '{field.field_id}' references parent "
                    f"'{field.parent_field_id}' which is not in the resolved "
                    "field set.",
                )

            parent_letter = get_column_letter(parent_col_idx)
            # The parent cell is on the same row as the child cell —
            # ${parent_letter}2 here means "the parent cell on the
            # first data row"; openpyxl will adjust as the validation
            # applies down the range.
            formula1 = build_substitute_formula(
                f"{parent_letter}2", layout.substitutions
            )
            dv = DataValidation(
                type="list", formula1=formula1, allow_blank=True
            )
            dv.error = f"Select a valid {field.display_name}"
            dv.errorTitle = "Invalid selection"
            dv.showErrorMessage = True
            dv.add(cell_range)
            data_ws.add_data_validation(dv)
            continue

        # ── Flat dropdown ──
        if field.lookup_name:
            lookup = lookups[field.lookup_name]
            range_name = f"{NAMED_RANGE_PREFIX}FLAT_{_safe_identifier(lookup.lookup_name)}"
            dv = DataValidation(
                type="list", formula1=f"={range_name}", allow_blank=True
            )
            dv.error = f"Select a valid {field.display_name}"
            dv.errorTitle = "Invalid selection"
            dv.showErrorMessage = True
            dv.add(cell_range)
            data_ws.add_data_validation(dv)
            continue

        # ── Type-and-format validations ──
        dv = _build_type_validation(field, cell_range)
        if dv is not None:
            data_ws.add_data_validation(dv)


def _build_type_validation(field: FieldDef,
                           cell_range: str) -> Optional[DataValidation]:
    fmt = field.data_format or {}

    if field.data_type == "number":
        dv_type = "whole" if fmt.get("integer") else "decimal"
        dv = DataValidation(
            type=dv_type,
            operator="between" if "min" in fmt and "max" in fmt else
                     ("greaterThanOrEqual" if "min" in fmt else
                      "lessThanOrEqual" if "max" in fmt else None),
            formula1=str(fmt["min"]) if "min" in fmt else None,
            formula2=str(fmt["max"]) if "max" in fmt and "min" in fmt else None,
            allow_blank=not field.required,
        )
        dv.error = f"{field.display_name} must be a number"
        dv.errorTitle = "Invalid value"
        dv.showErrorMessage = True
        dv.add(cell_range)
        return dv

    if field.data_type == "date":
        dv = DataValidation(
            type="date",
            operator="between" if "min" in fmt and "max" in fmt else None,
            formula1=fmt.get("min"),
            formula2=fmt.get("max"),
            allow_blank=not field.required,
        )
        dv.error = f"{field.display_name} must be a valid date"
        dv.errorTitle = "Invalid date"
        dv.showErrorMessage = True
        dv.add(cell_range)
        return dv

    if field.data_type == "text" and "max_length" in fmt:
        dv = DataValidation(
            type="textLength",
            operator="lessThanOrEqual",
            formula1=str(fmt["max_length"]),
            allow_blank=not field.required,
        )
        dv.error = f"{field.display_name} cannot exceed {fmt['max_length']} characters"
        dv.errorTitle = "Too long"
        dv.showErrorMessage = True
        dv.add(cell_range)
        return dv

    return None


# ──────────────────────────────────────────────────────────────────────
# Substages 8–9 — Final formatting and serialization
# ──────────────────────────────────────────────────────────────────────

def _serialize(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_filename(client_name: str, variant_name: Optional[str]) -> str:
    safe_client = _safe_identifier(client_name)
    safe_variant = _safe_identifier(variant_name) if variant_name else "base"
    stamp = datetime.utcnow().strftime("%Y%m%d")
    return f"{safe_client}_{safe_variant}_{stamp}.xlsx"


# ──────────────────────────────────────────────────────────────────────
# Errors
# ──────────────────────────────────────────────────────────────────────

class BuildError(Exception):
    """Structured error raised when a substage cannot proceed."""

    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


# ──────────────────────────────────────────────────────────────────────
# Top-level orchestration
# ──────────────────────────────────────────────────────────────────────

def build(build_input: BuildInput) -> BuildOutput:
    """
    Typed entry point. Composes the substages.
    """
    # 1. Resolve the field list (already variant-filtered by caller).
    fields = _resolve_fields(build_input)

    # 2. Resolve the lookups the resolved fields actually need.
    lookups = _resolve_lookups(fields, build_input.lookups)

    # 3. Lay out the reference sheet and derive the substitution map.
    layout = _lay_out_reference(lookups)

    # 4. Create the workbook (two sheets, reference hidden).
    wb, data_ws, ref_ws = _create_workbook()

    # 5. Header row on the data sheet.
    _write_data_header(data_ws, fields)

    # 6. Reference content (the lookup values), then register the
    # defined names so the formulas in step 7 have something to point at.
    _write_reference_content(ref_ws, lookups, layout, layout.substitutions)
    _register_defined_names(wb, layout)

    # 7. Data validation rules (flat dropdowns, dependent dropdowns,
    # type/format constraints).
    _apply_data_validations(data_ws, fields, lookups, layout)

    # 9. Serialize.
    workbook_bytes = _serialize(wb)

    # 10. Compose the output.
    return BuildOutput(
        file_content_base64=base64.b64encode(workbook_bytes).decode("ascii"),
        suggested_filename=_build_filename(
            build_input.client_name, build_input.variant_name
        ),
        metadata={
            "sheet_names": [DATA_SHEET_NAME, REFERENCE_SHEET_NAME],
            "byte_size": len(workbook_bytes),
            "field_count": len(fields),
            "row_count": 0,
            "data_row_capacity": DATA_ROWS,
        },
    )


# ──────────────────────────────────────────────────────────────────────
# Workato py_eval entry point
# ──────────────────────────────────────────────────────────────────────

def build_xlsx_template(payload: Dict[str, Any]) -> Dict[str, Any]:
      """
    Workato-runnable entry point. Accepts a plain dict; returns a plain
    dict. Wraps the typed `build` function.

    Expected payload shape:
        {
            "fields": [ {field_id, display_name, position, data_type,
                         required, lookup_name, parent_field_id,
                         data_format}, ... ],
            "lookups": { lookup_name: { "flat_values": [...] }
                         | { "parent_groups": { parent: [child, ...] } } },
            "client_name": str,
            "variant_name": str | None,
        }
    """
    build_input = _payload_to_input(payload)
    try:
        output = build(build_input)
    except BuildError as e:
        if e.code == "EMPTY_FIELD_LIST":
            return {
                "status": "empty_variant",
                "file_content": None,
                "suggested_filename": None,
                "metadata": None,
                "error": {
                    "code": "empty_variant",
                    "message": e.message,
                },
            }
        raise  # other BuildErrors propagate to the recipe's catch
    return {
        "status": "success",
        "file_content": output.file_content_base64,
        "suggested_filename": output.suggested_filename,
        "metadata": output.metadata,
        "error": None,
    }


def _payload_to_input(payload: Dict[str, Any]) -> BuildInput:
    fields = [
        FieldDef(
            field_id=f["field_id"],
            display_name=f["display_name"],
            position=int(f["position"]),
            data_type=f["data_type"],
            required=bool(f.get("required", False)),
            lookup_name=f.get("lookup_name") or None,
            parent_field_id=f.get("parent_field_id") or None,
            data_format=f.get("data_format") or {},
        )
        for f in payload.get("fields", [])
    ]

    lookups: Dict[str, LookupDef] = {}
    for name, body in (payload.get("lookups") or {}).items():
        lookups[name] = LookupDef(
            lookup_name=name,
            flat_values=body.get("flat_values"),
            parent_groups=body.get("parent_groups"),
        )

    return BuildInput(
        fields=fields,
        lookups=lookups,
        client_name=payload["client_name"],
        variant_name=payload.get("variant_name") or None,
    )
