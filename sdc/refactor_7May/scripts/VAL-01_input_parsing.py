"""
VAL-01 Step 9.2 — Parse XLSX submission

Drop-in replacement for the previous step. Changes from prior version:
- Output: 'rows' (array of object) → 'rows_json' (string)
- Happy path: json.dumps(rows, default=str) to produce a JSON string the
  connector's validate_upload can consume directly via its upload_data_json
  input. default=str handles openpyxl's native datetime/date values.
- Error paths: 'rows_json': '[]' (the JSON string for an empty array) for
  shape consistency; functionally these values are not consumed because
  in-band IFs route structural failures around the connector call.

Input schema (Workato py_eval step):
  - xlsx_base64: string (base64-encoded XLSX file contents)
  - canonical_model_json: string (the canonical model serialized as JSON)

Output schema (Workato py_eval step):
  - ok: boolean
  - error_code: string (nullable)
  - error_message: string (nullable)
  - rows_json: string (JSON-serialized array of row objects; "[]" on error)
  - row_count: integer

Failure modes (all return ok=False; recipe routes to structural_failure):
  - Could not decode inputs (base64 or canonical_model_json) → unparseable
  - openpyxl could not open XLSX → unparseable
  - Required sheet not found → structurally_invalid
  - Header row empty → structurally_invalid
  - Required headers missing from XLSX → structurally_invalid
"""

import base64
import json
from io import BytesIO


def main(input):
    # ── Decode inputs ──────────────────────────────────────────────────
    try:
        xlsx_bytes = base64.b64decode(input["xlsx_base64"])
        canonical_model = json.loads(input["canonical_model_json"])
    except Exception as e:
        return {
            "ok": False,
            "error_code": "err_submission_unparseable",
            "error_message": f"Could not decode inputs: {e}",
            "rows_json": "[]",
            "row_count": 0,
        }

    expected_sheet = canonical_model.get("_meta", {}).get("expected_sheet_name", "Data")
    field_names = [f["field_name"] for f in canonical_model.get("cfg_fields", [])]

    # ── Open workbook ──────────────────────────────────────────────────
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception as e:
        return {
            "ok": False,
            "error_code": "err_submission_unparseable",
            "error_message": f"Could not open XLSX: {e}",
            "rows_json": "[]",
            "row_count": 0,
        }

    # ── Locate expected sheet ──────────────────────────────────────────
    if expected_sheet not in wb.sheetnames:
        return {
            "ok": False,
            "error_code": "err_submission_structurally_invalid",
            "error_message": (
                f"Required sheet '{expected_sheet}' not found. "
                f"Found: {wb.sheetnames}"
            ),
            "rows_json": "[]",
            "row_count": 0,
        }

    ws = wb[expected_sheet]

    # ── Validate header row ────────────────────────────────────────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {
            "ok": False,
            "error_code": "err_submission_structurally_invalid",
            "error_message": "Header row is empty",
            "rows_json": "[]",
            "row_count": 0,
        }

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    missing_fields = [fn for fn in field_names if fn not in headers]
    if missing_fields:
        return {
            "ok": False,
            "error_code": "err_submission_structurally_invalid",
            "error_message": f"Missing required headers: {missing_fields}",
            "rows_json": "[]",
            "row_count": 0,
        }

    # ── Extract data rows ──────────────────────────────────────────────
    rows = []
    for row_tuple in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, cell in enumerate(row_tuple):
            if i < len(headers) and headers[i]:
                record[headers[i]] = cell
        # Skip blank rows (no non-blank values)
        if any(v is not None and str(v).strip() != "" for v in record.values()):
            rows.append(record)

    # ── Serialize for the connector ────────────────────────────────────
    # default=str handles datetime/date values from openpyxl (which would
    # otherwise raise TypeError: Object of type datetime is not JSON
    # serializable). The connector's check_data_type / check_data_format
    # already accept string-coerced values.
    rows_json = json.dumps(rows, default=str)

    return {
        "ok": True,
        "error_code": None,
        "error_message": None,
        "rows_json": rows_json,
        "row_count": len(rows),
    }



"""
VAL-01 Step 9.6 — Parse extracted.json (manual entry path)

Drop-in replacement for the previous step. Changes from prior version:
- Output: 'rows' (array of object) → 'rows_json' (string)
- Happy path: pass through the original JSON string (no parse-then-dump
  round trip). The input is already a JSON string from FileStorage; we
  just validate that it's a JSON array and forward.
- Error paths: 'rows_json': '[]' for shape consistency.

Input schema (Workato py_eval step):
  - content: string (the contents of the extracted.json file from
             FileStorage; expected to be a JSON-serialized array of row
             objects written by an upstream R3 step)

Output schema (Workato py_eval step):
  - ok: boolean
  - error_message: string (nullable)
  - rows_json: string (the original content if valid; "[]" on error)
  - row_count: integer

Failure modes (all return ok=False; recipe routes to verdict_status="error"):
  - content is not parseable as JSON → unparseable
  - content parses but is not a JSON array → wrong shape

Note: the recipe's structural-failure path is intended for XLSX-side
problems (corrupt files, missing headers). Manual entry coming through
malformed extracted.json indicates an R3 upstream bug, not a supplier
problem, so verdict_status="error" is the correct routing.
"""

import json


def main(input):
    content = input["content"]

    # ── Parse and validate shape ───────────────────────────────────────
    try:
        parsed = json.loads(content)
    except Exception as e:
        return {
            "ok": False,
            "error_message": f"Could not parse extracted.json: {e}",
            "rows_json": "[]",
            "row_count": 0,
        }

    if not isinstance(parsed, list):
        return {
            "ok": False,
            "error_message": "extracted.json is not a JSON array",
            "rows_json": "[]",
            "row_count": 0,
        }

    # ── Pass through the original string ───────────────────────────────
    # Avoids a parse-then-dump round trip. The source is already JSON, so
    # there are no datetime serialization concerns; the validated `parsed`
    # is used only for row_count.
    return {
        "ok": True,
        "error_message": None,
        "rows_json": content,
        "row_count": len(parsed),
    }
